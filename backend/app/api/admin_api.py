from fastapi import APIRouter, HTTPException
from app.config import get_connection
from app.utils.auth_utils import check_admin
import sqlite3
import pandas as pd
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import Font
router = APIRouter(prefix="/admin", tags=["Admin"])


# ===============================
# Lấy danh sách người dùng
# ===============================
@router.get("/users")
def get_users(role: str):

    check_admin(role)

    conn = get_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT 
                nd.MaNguoiDung,
                nd.HoTen,
                nd.Email,
                vt.TenVaiTro,
                tk.TrangThai  
            FROM NguoiDung nd
            JOIN TaiKhoan tk ON nd.MaNguoiDung = tk.MaNguoiDung
            JOIN VaiTro vt ON tk.MaVaiTro = vt.MaVaiTro
        """)

        users = [
            {
                "id": row[0],
                "name": row[1],
                "email": row[2],
                "role": row[3],
                "status": row[4]
            }
            for row in cursor.fetchall()
        ]

        return users

    finally:
        cursor.close()
        conn.close()


# ===============================
# Lịch sử chat
# ===============================
@router.get("/chat-history")
def chat_history(role: str):

    check_admin(role)

    conn = get_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT 
                nd.HoTen,
                lc.CauHoi,
                lc.TraLoi,
                lc.ThoiGian
            FROM LichSuChat lc
            JOIN NguoiDung nd 
            ON lc.MaNguoiDung = nd.MaNguoiDung
            ORDER BY lc.ThoiGian DESC
        """)

        history = [
            {
                "user": row[0],
                "question": row[1],
                "answer": row[2],
                "time": str(row[3])
            }
            for row in cursor.fetchall()
        ]

        return history

    finally:
        cursor.close()
        conn.close()


# ===============================
# Thống kê hệ thống
# ===============================
@router.get("/stats")
def stats():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM NguoiDung")
    total_users = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM LichSuChat")
    total_chats = cursor.fetchone()[0]

    cursor.execute("""
        SELECT AVG(CAST(DoChinhXac AS FLOAT))
        FROM DanhGiaChatbot
        WHERE DoChinhXac IS NOT NULL
    """)
    accuracy = cursor.fetchone()[0] or 0

    conn.close()

    return {
        "total_users": total_users,
        "total_chats": total_chats,
        "accuracy": round(accuracy,2)
    }
@router.get("/report-preview")
def report_preview(role: str, start: str, end: str):

    check_admin(role)

    conn = get_connection()
    cursor = conn.cursor() if conn else None
    if not cursor: raise HTTPException(status_code=500, detail="Database connection failed")

    # ========================
    # Tổng user
    # ========================
    cursor.execute("SELECT COUNT(*) FROM NguoiDung")
    total_users = cursor.fetchone()[0]

    # ========================
    # Tổng câu hỏi
    # ========================
    cursor.execute("""
        SELECT COUNT(*) 
        FROM LichSuChat
        WHERE CAST(ThoiGian AS DATE) BETWEEN ? AND ?
    """, (start, end))
    total_questions = cursor.fetchone()[0]

    # ========================
    # User hoạt động
    # ========================
    cursor.execute("""
        SELECT COUNT(DISTINCT MaNguoiDung)
        FROM LichSuChat
        WHERE CAST(ThoiGian AS DATE) BETWEEN ? AND ?
    """, (start, end))
    active_users = cursor.fetchone()[0]

    # ========================
    # TOP USER
    # ========================
    cursor.execute("""
        SELECT TOP 5 N.HoTen, COUNT(*) as total
        FROM LichSuChat C
        JOIN NguoiDung N ON C.MaNguoiDung = N.MaNguoiDung
        WHERE CAST(C.ThoiGian AS DATE) BETWEEN ? AND ?
        GROUP BY N.HoTen
        ORDER BY total DESC
    """, (start, end))

    top_users = [
        {"name": row[0], "total": row[1]}
        for row in cursor.fetchall()
    ]

    # ========================
    # TOP CHỦ ĐỀ (phân tích câu hỏi)
    # ========================
    cursor.execute("""
        SELECT CauHoi
        FROM LichSuChat
        WHERE CAST(ThoiGian AS DATE) BETWEEN ? AND ?
    """, (start, end))

    questions = [row[0] for row in cursor.fetchall()]

    topics = {}

    for q in questions:

        q = str(q).lower()

        if "ly hôn" in q:
            topic = "Ly hôn"

        elif "nuôi con" in q:
            topic = "Quyền nuôi con"

        elif "tài sản" in q:
            topic = "Chia tài sản"

        else:
            topic = "Khác"

        topics[topic] = topics.get(topic, 0) + 1

    # sắp xếp top chủ đề
    sorted_topics = sorted(topics.items(), key=lambda x: x[1], reverse=True)

    top_topics = [
        {"name": topic, "total": count}
        for topic, count in sorted_topics[:5]
    ]

    conn.close()

    return {
        "total_users": total_users,
        "total_questions": total_questions,
        "active_users": active_users,
        "top_users": top_users,
        "top_topics": top_topics
    }
from fastapi.responses import FileResponse
import pandas as pd
import pyodbc

@router.get("/export-report")
def export_report(start: str, end: str):

    conn = get_connection()
    cursor = conn.cursor() if conn else None
    if not cursor: raise HTTPException(status_code=500, detail="Database connection failed")

    # =========================
    # Tổng người dùng
    # =========================
    cursor.execute("SELECT COUNT(*) FROM NguoiDung")
    total_users = cursor.fetchone()[0]

    # =========================
    # Tổng câu hỏi
    # =========================
    cursor.execute("""
        SELECT COUNT(*)
        FROM LichSuChat
        WHERE CAST(ThoiGian AS DATE) BETWEEN ? AND ?
    """, (start, end))

    total_questions = cursor.fetchone()[0]

    # =========================
    # User hoạt động
    # =========================
    cursor.execute("""
        SELECT COUNT(DISTINCT MaNguoiDung)
        FROM LichSuChat
        WHERE CAST(ThoiGian AS DATE) BETWEEN ? AND ?
    """, (start, end))

    active_users = cursor.fetchone()[0]

    # =========================
    # TOP USER
    # =========================
    cursor.execute("""
        SELECT TOP 5 nd.HoTen, COUNT(*) as SoCauHoi
        FROM LichSuChat lsc
        JOIN NguoiDung nd ON nd.MaNguoiDung = lsc.MaNguoiDung
        WHERE CAST(lsc.ThoiGian AS DATE) BETWEEN ? AND ?
        GROUP BY nd.HoTen
        ORDER BY SoCauHoi DESC
    """, (start, end))

    top_users = cursor.fetchall()

    # =========================
    # Lấy dữ liệu chi tiết chat
    # =========================
    query = """
    SELECT 
        MaChat,
        MaNguoiDung,
        CauHoi,
        TraLoi,
        ThoiGian
    FROM LichSuChat
    WHERE CAST(ThoiGian AS DATE) BETWEEN ? AND ?
    """

    df = pd.read_sql(query, conn, params=[start, end])

    # =========================
    # Thống kê theo ngày
    # =========================
    cursor.execute("""
        SELECT 
            CAST(ThoiGian AS DATE) AS Ngay,
            COUNT(*) AS SoCau
        FROM LichSuChat
        WHERE ThoiGian BETWEEN ? AND ?
        GROUP BY CAST(ThoiGian AS DATE)
        ORDER BY Ngay
        """, (start, end))
    daily_stats = cursor.fetchall()
    daily_stats = [(row[0], row[1]) for row in daily_stats]
    df_daily = pd.DataFrame(daily_stats, columns=["Ngay", "SoCau"])

    # =========================
    # Phân tích chủ đề
    # =========================
    topics = {}

    for q in df["CauHoi"]:
        q = str(q).lower()

        if "ly hôn" in q:
            topics["Ly hôn"] = topics.get("Ly hôn", 0) + 1

        elif "nuôi con" in q:
            topics["Quyền nuôi con"] = topics.get("Quyền nuôi con", 0) + 1

        elif "tài sản" in q:
            topics["Chia tài sản"] = topics.get("Chia tài sản", 0) + 1

        else:
            topics["Khác"] = topics.get("Khác", 0) + 1

    sorted_topics = sorted(topics.items(), key=lambda x: x[1], reverse=True)

    # =========================
    # Tạo Excel
    # =========================
    file_name = "bao_cao_chatbot.xlsx"

    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:

        # =========================
        # Sheet Dashboard
        # =========================
        wb = writer.book
        ws = wb.create_sheet("BaoCao")

        ws["A1"] = "BÁO CÁO HOẠT ĐỘNG CHATBOT AI"
        ws["A1"].font = Font(size=16, bold=True)

        ws["A3"] = f"Từ ngày: {start}"
        ws["A4"] = f"Đến ngày: {end}"

        ws["A6"] = "Tổng người dùng"
        ws["B6"] = total_users

        ws["A7"] = "Tổng câu hỏi"
        ws["B7"] = total_questions

        ws["A8"] = "User hoạt động"
        ws["B8"] = active_users

        # =========================
        # TOP USER
        # =========================
        ws["A10"] = "TOP USER"

        row = 11
        for i, user in enumerate(top_users, start=1):

            ws[f"A{row}"] = user[0]
            ws[f"B{row}"] = user[1]

            row += 1

        # =========================
        # TOP CHỦ ĐỀ
        # =========================
        ws["D10"] = "TOP CHỦ ĐỀ"

        row = 11
        for topic, count in sorted_topics[:5]:

            ws[f"D{row}"] = topic
            ws[f"E{row}"] = count

            row += 1

        # =========================
        # Biểu đồ TOP USER
        # =========================
        chart = BarChart()
        chart.title = "Top User hỏi nhiều nhất"

        data = Reference(ws, min_col=2, min_row=11, max_row=15)
        cats = Reference(ws, min_col=1, min_row=11, max_row=15)

        chart.add_data(data)
        chart.set_categories(cats)

        ws.add_chart(chart, "A16")

        # =========================
        # Pie chart TOP CHỦ ĐỀ
        # =========================
        pie = PieChart()

        pie.title = "Phân bố chủ đề câu hỏi"

        data = Reference(ws, min_col=5, min_row=11, max_row=15)
        cats = Reference(ws, min_col=4, min_row=11, max_row=15)

        pie.add_data(data)
        pie.set_categories(cats)

        ws.add_chart(pie, "D16")

        # =========================
        # Sheet Chi tiết chat
        # =========================
        df.to_excel(writer, sheet_name="ChiTietChat", index=False)

        # =========================
        # Sheet thống kê ngày
        # =========================
        df_daily.to_excel(writer, sheet_name="ThongKeNgay", index=False)

        ws_daily = writer.book["ThongKeNgay"]

        line = LineChart()
        line.title = "Số câu hỏi theo ngày"

        data = Reference(ws_daily, min_col=2, min_row=1, max_row=len(df_daily) + 1)
        cats = Reference(ws_daily, min_col=1, min_row=2, max_row=len(df_daily) + 1)

        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)

        ws_daily.add_chart(line, "E2")

    conn.close()

    return FileResponse(
        file_name,
        filename="bao_cao_chatbot.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# ===============================
# Khóa tài khoản
# ===============================
@router.put("/lock-user/{user_id}")
def lock_user(user_id: int, role: str):

    check_admin(role)

    conn = get_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            UPDATE TaiKhoan
            SET TrangThai = N'Đã khóa'
            WHERE MaNguoiDung = ?
        """, (user_id,))

        conn.commit()

        return {"message": "Tài khoản đã bị khóa"}

    finally:
        cursor.close()
        conn.close()
# ===============================
# Mở khóa tài khoản
# ===============================
@router.put("/unlock-user/{user_id}")
def unlock_user(user_id: int, role: str):

    check_admin(role)

    conn = get_connection()
    cursor = conn.cursor()

    try:

        cursor.execute("""
            UPDATE TaiKhoan
            SET TrangThai = N'Hoạt động'
            WHERE MaNguoiDung = ?
        """, (user_id,))

        conn.commit()

        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Không tìm thấy tài khoản")

        return {"message": "Tài khoản đã được mở khóa"}

    finally:
        cursor.close()
        conn.close()
@router.get("/chart-data")
def chart_data():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT CAST(ThoiGian AS DATE), COUNT(*)
        FROM LichSuChat
        GROUP BY CAST(ThoiGian AS DATE)
        ORDER BY CAST(ThoiGian AS DATE)
    """)

    rows = cursor.fetchall()

    labels = []
    values = []

    for r in rows:
        labels.append(str(r[0]))
        values.append(r[1])

    return {
        "labels": labels,
        "values": values
    }
@router.get("/topic-chart")
def topic_chart():
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT TOP 5 
            CASE 
                WHEN CauHoi LIKE N'%kết hôn%' THEN N'Kết hôn'
                WHEN CauHoi LIKE N'%ly hôn%' THEN N'Ly hôn'
                WHEN CauHoi LIKE N'%tài sản%' THEN N'Tài sản'
                WHEN CauHoi LIKE N'%con%' THEN N'Quyền nuôi con'
                ELSE N'Khác'
            END AS Topic,
            COUNT(*) as Total
        FROM LichSuChat
        GROUP BY 
            CASE 
                WHEN CauHoi LIKE N'%kết hôn%' THEN N'Kết hôn'
                WHEN CauHoi LIKE N'%ly hôn%' THEN N'Ly hôn'
                WHEN CauHoi LIKE N'%tài sản%' THEN N'Tài sản'
                WHEN CauHoi LIKE N'%con%' THEN N'Quyền nuôi con'
                ELSE N'Khác'
            END
        ORDER BY Total DESC
    """)

    rows = cursor.fetchall()

    labels = []
    values = []

    for r in rows:
        labels.append(r.Topic)
        values.append(r.Total)

    return {
        "labels": labels,
        "values": values
    }
@router.get("/rating-stats")
def rating_stats():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT DiemDanhGia, COUNT(*)
        FROM DanhGiaChatbot
        GROUP BY DiemDanhGia
        ORDER BY DiemDanhGia
    """)

    rows = cursor.fetchall()

    conn.close()

    return {
        "labels": [r[0] for r in rows],
        "values": [r[1] for r in rows]
    }
@router.get("/ai-accuracy")
def get_ai_accuracy(role: str):
    check_admin(role)
    conn = get_connection()
    cursor = conn.cursor()
    try:
        # Tính trung bình cộng cột DiemDanhGia
        cursor.execute("SELECT AVG(CAST(DiemDanhGia AS FLOAT)) FROM LichSuChat WHERE DiemDanhGia IS NOT NULL")
        result = cursor.fetchone()[0]
        return {"accuracy": round(result, 1) if result else 0}
    finally:
        conn.close()
@router.get("/accuracy")
def get_accuracy():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT AVG(CAST(DoChinhXac AS FLOAT))
        FROM DanhGiaChatbot
        WHERE DoChinhXac IS NOT NULL
    """)

    avg = cursor.fetchone()[0]

    conn.close()

    return {
        "accuracy": round(avg or 0, 2)
    }
@router.get("/thongke-danhgia")
def thongke_danhgia():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT ROUND(AVG(DoChinhXac),2) 
    FROM DanhGiaChatbot
    """)

    avg_accuracy = cursor.fetchone()[0] or 0

    cursor.execute("""
    SELECT DiemDanhGia, COUNT(*)
    FROM DanhGiaChatbot
    GROUP BY DiemDanhGia
    """)

    rating_data = cursor.fetchall()

    result = {
        "accuracy": avg_accuracy,
        "ratings": [
            {"star": row[0], "count": row[1]} for row in rating_data
        ]
    }

    conn.close()

    return result
